from io import BytesIO
from pathlib import Path
from typing import Union
from .cfdi import FacturaFiscal, Concepto,Emisor,Receptor


def convertir_facturas_zip(file: Union[str, Path, BytesIO] = None):

    from zipfile import ZipFile

    if file is None:
        return "Debe de enviar un archivo o ruta del archivo."

    if not isinstance(file, BytesIO):
        if not file.exists():
            return "El archivo en la ruta {} no existe".format(file)
        
        file_read = open(str(file), "rb")
        file = BytesIO(file_read.read())
        file_read.close()

    facturas: list[FacturaFiscal] = []

    with ZipFile(file, "r") as zip_file:
        for file_name in zip_file.namelist():
            if file_name.lower().endswith(".xml"):
                with zip_file.open(file_name) as xml_content:
                    factura = FacturaFiscal.parse_from_xml(xml_content.read())
                    if factura is not None:
                        facturas.append(factura)

    return sorted([f for f in facturas if f.fecha], key=lambda x : x.fecha)


def exportar_facturas_excel(facturas: list[FacturaFiscal]) -> str:

    from openpyxl import Workbook
    from datetime import datetime

    class EVConcepto(Concepto):
        emisor: Emisor
        receptor: Receptor
        uuid: str
        fecha: datetime | None

    list_conceptos: list[EVConcepto] = []
    for factura in facturas:
        for concepto in factura.conceptos:
            c = EVConcepto(**concepto.__dict__)
            c.fecha = factura.fecha
            c.uuid = factura.uuid
            c.emisor = factura.emisor
            c.receptor = factura.receptor
            list_conceptos.append(c)

    wb = Workbook()

    sheetname = 'MovFact'
    if sheetname not in wb.sheetnames:
        wb.create_sheet(sheetname)

    ws1 = wb.worksheets[0]
    ws1.title = 'Facturas'

    headers = [
        "Periodo",
        "Periodo Declarado",
        "Fecha",
        "Uuid",
        "RFC Emisor",
        "Emisor",
        "RFC Receptor",
        "Receptor",
        "Subtotal",
        "IVA Trasladado",
        "ISR Retenido",
        "Total"
    ]

    for i, header in enumerate(headers):
        ws1.cell(1, i + 1, header.upper())

    ws2 = wb[sheetname]
    for i, header in enumerate(headers):
        ws2.cell(1, i + 1, header.upper())

    
    for i, fact in enumerate(facturas):
        i += 2

        ws1.cell(i, 1, "=MONTH(C{})".format(i))
        ws1.cell(i, 2, "=MONTH(C{})".format(i))
        ws1.cell(i, 3, fact.fecha)
        ws1.cell(i, 4, fact.uuid)
        ws1.cell(i, 5, fact.emisor.rfc)
        ws1.cell(i, 6, fact.emisor.nombre)
        ws1.cell(i, 7, fact.receptor.rfc)
        ws1.cell(i, 8, fact.receptor.nombre)
        ws1.cell(i, 9, fact.subtotal)

        for impuesto in fact.impuestos:
            if impuesto.tipo == 'traslado':
                if impuesto.impuesto == 2:
                    ws1.cell(i, 10, impuesto.importe)
            else:
                ws1.cell(i, 11, impuesto.importe * -1)
        
        ws1.cell(i, 12, "=SUM(I{}:K{})".format(i,i))

    accounting_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
    for row in ws1.iter_rows(min_col=9, max_col=12):
        for cell in row:
            cell.number_format = accounting_format

    
    for i, concepto in enumerate(list_conceptos):
        i += 2
        ws2.cell(i, 1, "=MONTH(C{})".format(i))
        ws2.cell(i, 2, "=MONTH(C{})".format(i))
        ws2.cell(i, 3, concepto.fecha)
        ws2.cell(i, 4, concepto.uuid)
        ws2.cell(i, 5, concepto.emisor.rfc)
        ws2.cell(i, 6, concepto.emisor.nombre)
        ws2.cell(i, 7, concepto.receptor.rfc)
        ws2.cell(i, 8, concepto.receptor.nombre)
        ws2.cell(i, 9, concepto.importe)

        for impuesto in concepto.impuestos:
            if impuesto.tipo == 'traslado':
                if impuesto.impuesto == 2:
                    ws2.cell(i, 10, impuesto.importe)
            else:
                ws2.cell(i, 11, impuesto.importe * -1)
        
        ws2.cell(i, 12, "=SUM(I{}:K{})".format(i,i))



    accounting_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
    for row in ws2.iter_rows(min_col=9, max_col=12):
        for cell in row:
            cell.number_format = accounting_format


    wb.save('RelacionCFDI.xlsx')

    return None
